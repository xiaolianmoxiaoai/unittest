# -*- coding: utf-8 -*-
"""
-------------------------------------------------
   File Name：     get_data
   Description :  从excel中读取数据
   Author :       WWQ
   date：          2020/3/28
-------------------------------------------------
   Change Activity:
                   2020/3/28:
-------------------------------------------------
"""
__author__ = 'WWQ'
from openpyxl import load_workbook
from collections import namedtuple

class GetExcelDatas:
    '''
    读取excel中的测试数据和测试用例
    '''
    def __init__(self):
        pass
    def __call__(self, file, sheet_name):
        ws = load_workbook(file)  # 加载excel文件
        sheet = ws.get_sheet_by_name(sheet_name)  # 通过sheet页名称读取
        sheet_head_tuple = tuple(sheet.iter_rows(max_row=1, values_only=True))[0]  # 获取表格第一行标题
        case_list = []
        Cases = namedtuple("Cases", sheet_head_tuple, rename=True)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            case_list.append(Cases(*row))
        return case_list

if __name__ == '__main__':
    data = GetExcelDatas()
    datas = data(file="./测试用例.xlsx", sheet_name="四则运算")
    print(datas[2].测试序号)
    pass



